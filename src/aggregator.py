import pandas as pd
import os
import re
import sys
import subprocess
import json
import warnings
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import xlrd # Explicit import for PyInstaller hidden import
import unicodedata

# Suppress openpyxl warnings if any
warnings.filterwarnings("ignore")

# When packaged as a --onefile exe, all bundled files land in sys._MEIPASS.
# In that case __file__ points inside the temp extraction dir, so the '..'
# relative path would escape it.  We therefore use _MEIPASS directly.
if getattr(sys, 'frozen', False):
    _XLRD_LEGACY_DIR = os.path.join(sys._MEIPASS, 'xlrd_legacy')
else:
    _XLRD_LEGACY_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'xlrd_legacy')

_XLS_FMT_SCRIPT = '''
import sys, json
legacy_dir = sys.argv[1]
xls_path   = sys.argv[2]
nrows      = int(sys.argv[3])
ncols      = int(sys.argv[4])
sys.path.insert(0, legacy_dir)

import xlrd

BORDER_MAP = {0:None,1:'thin',2:'medium',3:'dashed',4:'dotted',5:'thick',
              6:'double',7:'hair',8:'mediumDashed',9:'dashDot',
              10:'mediumDashDot',11:'dashDotDot',12:'mediumDashDotDot',
              13:'slantDashDot'}
HOR_MAP  = {0:'general',1:'left',2:'center',3:'right',4:'fill',
            5:'justify',6:'centerContinuous',7:'distributed'}
VERT_MAP = {0:'top',1:'center',2:'bottom',3:'justify',4:'distributed'}

def rgb_hex(rgb):
    if rgb is None: return None
    return 'FF{:02X}{:02X}{:02X}'.format(rgb[0], rgb[1], rgb[2])

def side_d(wb, style, cidx):
    s = BORDER_MAP.get(style)
    if not s: return None
    rgb = wb.colour_map.get(cidx)
    color = rgb_hex(rgb) if rgb else 'FF000000'
    return {'s': s, 'c': color}

try:
    wb = xlrd.open_workbook(xls_path, formatting_info=True, encoding_override='cp932')
except Exception:
    wb = xlrd.open_workbook(xls_path, formatting_info=True)
ws = wb.sheet_by_index(0)

nrows = min(nrows, ws.nrows)
ncols = min(ncols, ws.ncols)

out = {'rh': {}, 'cw': {}, 'mg': [], 'xf': {}, 'cl': []}

for r in range(ws.nrows):
    ri = ws.rowinfo_map.get(r)
    if ri:
        out['rh'][str(r)] = round(ri.height / 20.0, 2)

for c in range(ws.ncols):
    ci = ws.colinfo_map.get(c)
    if ci:
        out['cw'][str(c)] = round(ci.width / 256.0, 3)

for mc in ws.merged_cells:
    rlo, rhi, clo, chi = mc
    if rlo < nrows and clo < ncols:
        out['mg'].append([rlo, rhi, clo, chi])

xf_cache = {}
cells = [[None] * ncols for _ in range(nrows)]
for r in range(nrows):
    for c in range(ncols):
        idx = ws.cell_xf_index(r, c)
        if idx not in xf_cache:
            xf = wb.xf_list[idx]
            bg_idx = xf.background.pattern_colour_index
            if bg_idx not in (64, 65, 0x7FFF):
                bg_rgb = wb.colour_map.get(bg_idx)
                bg = rgb_hex(bg_rgb)
            else:
                bg = None
            f = wb.font_list[xf.font_index]
            b = xf.border
            al = xf.alignment
            xf_cache[idx] = {
                'bg': bg,
                'bo': bool(f.bold),
                'fn': f.name,
                'fs': round(f.height / 20.0, 1),
                'bl': side_d(wb, b.left_line_style,  b.left_colour_index),
                'br': side_d(wb, b.right_line_style, b.right_colour_index),
                'bt': side_d(wb, b.top_line_style,   b.top_colour_index),
                'bb': side_d(wb, b.bottom_line_style, b.bottom_colour_index),
                'ha': HOR_MAP.get(al.hor_align, 'general'),
                'va': VERT_MAP.get(al.vert_align, 'bottom'),
                'wr': bool(al.text_wrapped),
            }
        cells[r][c] = idx

out['xf'] = {str(k): v for k, v in xf_cache.items()}
out['cl'] = cells
print(json.dumps(out))
'''

# ==============================================================================
# CONFIGURATION (設定)
# もし大学のシステムが変わり、Excelの「列」や「行」の位置が変わった場合は、以下の数字を調整してください
# If Excel format changes, please modify these numbers
# ==============================================================================
CONFIG = {
    # --- 1. Comment Sheet (コメントシート) ---
    # A=0, B=1, C=2, D=3, E=4, F=5, G=6 ...
    "COL_SUB_ID": 0,    # A列: Submission ID
    "COL_COURSE": 2,    # C列: Course Name (Year checks this)
    "COL_NAME": 4,      # E列: Student Name
    "COL_ID": 5,        # F列: Student ID
    "COL_COMMENT": 6,   # G列: Comment Body

    # Minimum columns required to be valid
    # precise logic: max(indices) + 1. Since COL_COMMENT is 6, we need 7.
    "MIN_COLS": 7,

    # --- 2. Attendance Sheet (出席簿) ---
    "ATT_SKIP_ROWS": 6, # Number of header rows to skip (Data starts at Row 7)
    "ATT_COL_ID": 1,    # B列: Student ID (0-indexed)
    "ATT_COL_NAME": 2,  # C列: Student Name (0-indexed)
}
# ==============================================================================

# Original default values — used by the GUI "Reset to Defaults" button.
_DEFAULT_CONFIG = dict(CONFIG)


def load_config(filepath):
    """
    Load user-saved CONFIG overrides from a JSON file and apply them in-place.
    Called once at GUI startup.  Unknown / non-integer keys are silently ignored
    so that a config file written by an older version stays compatible.
    """
    if not os.path.isfile(filepath):
        return
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            user_cfg = json.load(f)
        for key in _DEFAULT_CONFIG:          # only known keys
            if key in user_cfg and isinstance(user_cfg[key], int):
                CONFIG[key] = user_cfg[key]
        print(f"User config loaded from {os.path.basename(filepath)}")
    except Exception as e:
        print(f"Warning: could not load user config: {e}")


def save_config(filepath, cfg=None):
    """
    Persist CONFIG (or a given dict) to a JSON file next to the executable.
    """
    data = cfg if cfg is not None else dict(CONFIG)
    try:
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"User config saved to {os.path.basename(filepath)}")
    except Exception as e:
        print(f"Warning: could not save user config: {e}")


# Sentinel values that indicate a header/metadata row rather than a real student
_HEADER_IDS = {"学籍番号", "id", "student id", "headerid", "number"}


def _read_xls_fmt_store(xls_path, nrows, ncols):
    """
    Run xlrd 1.x in a subprocess to extract .xls cell formatting.
    Returns a dict with keys: 'rh' (row heights), 'cw' (col widths),
    'mg' (merged cells), 'xf' (format dicts), 'cl' (cell->xf index matrix).
    Returns None if the legacy xlrd is unavailable or the read fails.
    """
    if not os.path.isdir(_XLRD_LEGACY_DIR):
        print("Warning: xlrd_legacy not found; .xls formatting will not be copied.")
        return None
    try:
        result = subprocess.run(
            [sys.executable, '-c', _XLS_FMT_SCRIPT,
             _XLRD_LEGACY_DIR, xls_path, str(nrows), str(ncols)],
            capture_output=True, text=True, timeout=60,
            encoding='utf-8',
        )
        if result.returncode == 0 and result.stdout.strip():
            return json.loads(result.stdout)
        if result.stderr:
            print(f"Warning: xls format read: {result.stderr[:300]}")
    except Exception as e:
        print(f"Warning: could not read .xls formatting: {e}")
    return None


def _apply_att_cell_fmt(cell, fmt, caches):
    """Apply one cell's XLS formatting (fill/font/border/alignment) to an openpyxl cell."""
    fill_cache, font_cache, border_cache, align_cache = caches

    # Background fill
    bg = fmt.get('bg')
    if bg:
        if bg not in fill_cache:
            fill_cache[bg] = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
        cell.fill = fill_cache[bg]

    # Font
    bold      = fmt.get('bo', False)
    font_name = fmt.get('fn', 'Calibri')
    font_size = fmt.get('fs', 11.0)
    fkey = (bold, font_name, font_size)
    if fkey not in font_cache:
        font_cache[fkey] = Font(name=font_name, size=font_size, bold=bold)
    cell.font = font_cache[fkey]

    # Border
    def _side(sd):
        if sd is None:
            return Side(border_style=None)
        return Side(border_style=sd['s'], color=sd.get('c', 'FF000000'))

    bl, br, bt, bb = fmt.get('bl'), fmt.get('br'), fmt.get('bt'), fmt.get('bb')
    bkey = (str(bl), str(br), str(bt), str(bb))
    if bkey not in border_cache:
        border_cache[bkey] = Border(
            left=_side(bl), right=_side(br),
            top=_side(bt),  bottom=_side(bb),
        )
    cell.border = border_cache[bkey]

    # Alignment
    ha   = fmt.get('ha', 'general')
    va   = fmt.get('va', 'bottom')
    wrap = fmt.get('wr', False)
    akey = (ha, va, wrap)
    if akey not in align_cache:
        align_cache[akey] = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    cell.alignment = align_cache[akey]


# Module-level helper — defined once, not recreated per row
def _get_cell_val(cell):
    return str(cell.value) if cell.value is not None else ""


def _read_attendance_rows(filepath):
    """
    Read all rows from an attendance file (.xls or .xlsx) as a list of lists.

    Works around the old .xls format that openpyxl cannot open:
    - .xls  → xlrd (handles legacy formats including BIFF5/BIFF8)
    - .xlsx → openpyxl

    Each row is a Python list; empty cells are represented as None.
    Numeric cell values that are whole numbers are returned as int to avoid
    "21Z1007C.0"-style ID artefacts.
    """
    ext = os.path.splitext(filepath)[1].lower()

    if ext == '.xls':
        # Japanese .xls files are typically Shift-JIS (CP932).
        # The encoding_override only affects byte-string cells; UTF-16 cells
        # (common in BIFF8 / Excel 97+) are unaffected, so this is always safe.
        try:
            wb = xlrd.open_workbook(filepath, encoding_override='cp932')
        except Exception:
            wb = xlrd.open_workbook(filepath)  # fallback without override
        ws = wb.sheet_by_index(0)

        rows = []
        for r in range(ws.nrows):
            row = []
            for c in range(ws.ncols):
                ctype = ws.cell_type(r, c)
                val   = ws.cell_value(r, c)
                if ctype == xlrd.XL_CELL_EMPTY:
                    row.append(None)
                elif ctype == xlrd.XL_CELL_NUMBER:
                    # Store integers without .0 suffix (avoids "123456.0" IDs)
                    row.append(int(val) if val == int(val) else val)
                else:
                    row.append(val)
            rows.append(row)
        return rows

    else:  # .xlsx
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        return [[cell.value for cell in row] for row in ws.iter_rows()]


def _safe_str(val):
    """Convert a cell value to a stripped string; returns '' for None/empty."""
    if val is None:
        return ''
    return str(val).strip()


def process_files(input_files, output_file, target_year=None, attendance_file=None):
    """
    Reads selected Excel files, aggregates comments, and saves to output_file.

    - If attendance_file is provided: produces a combined workbook with the
      attendance sheet on the left and comment columns appended on the right,
      aligned to each student row.  Supports both .xls and .xlsx attendance
      files (the output is always a modern .xlsx file).
    - If attendance_file is omitted: produces a standalone comment pivot table.

    Filters by target_year if provided (checks Column C of comment sheets).
    """
    all_data = []

    print(f"Processing {len(input_files)} files...")
    print(f"Target Year Filter: {target_year if target_year else 'None'}")

    attendance_ids_order = []          # Ordered list of NormIDs from attendance sheet
    attendance_map_name = {}           # NormID -> display name (Kanji)
    attendance_map_original_id = {}    # NormID -> original ID string (for display)
    attendance_name_map = {}           # Normalised name -> NormID (for fuzzy fallback)
    att_rows_all = None                # All raw rows from attendance file (list of lists)

    if attendance_file:
        print(f"Loading attendance sheet: {os.path.basename(attendance_file)}")
        try:
            att_rows_all = _read_attendance_rows(attendance_file)

            skip_count = CONFIG["ATT_SKIP_ROWS"]
            id_0  = CONFIG["ATT_COL_ID"]
            name_0 = CONFIG["ATT_COL_NAME"]

            if len(att_rows_all) <= skip_count:
                return False, "Attendance sheet is too short/empty."

            for row in att_rows_all[skip_count:]:
                if len(row) <= id_0 or row[id_0] is None:
                    continue

                s_id   = _safe_str(row[id_0])
                s_name = _safe_str(row[name_0]) if len(row) > name_0 else ''

                if not s_id or s_id.lower() in _HEADER_IDS:
                    continue

                norm_id = s_id.lower()
                attendance_ids_order.append(norm_id)
                attendance_map_name[norm_id] = s_name
                attendance_map_original_id[norm_id] = s_id

                norm_name = re.sub(r'\s+', '', unicodedata.normalize('NFKC', s_name))
                if norm_name:
                    attendance_name_map[norm_name] = norm_id

            print(f"Loaded {len(attendance_ids_order)} students from attendance sheet.")

        except Exception as e:
            print(f"Error reading attendance sheet: {e}")
            return False, f"Error reading attendance sheet: {e}"

    # Read .xls formatting for faithful reproduction in output
    fmt_store = None
    if attendance_file and os.path.splitext(attendance_file)[1].lower() == '.xls':
        att_nrows = len(att_rows_all)
        att_ncols = max((len(r) for r in att_rows_all), default=0)
        print("Reading .xls formatting info...")
        fmt_store = _read_xls_fmt_store(attendance_file, att_nrows, att_ncols)
        if fmt_store:
            print("Formatting info loaded successfully.")

    # ──────────────────────────────────────────────────────────────────────────
    # Process comment files
    # ──────────────────────────────────────────────────────────────────────────
    for file_path in input_files:
        try:
            filename = os.path.basename(file_path)
            date_match = re.search(r'\d{4}-\d{2}-\d{2}', filename)
            date_str = date_match.group(0) if date_match else filename

            wb_in = load_workbook(file_path, data_only=True)
            ws_in = wb_in.active
            min_cols = CONFIG["MIN_COLS"]

            for row in ws_in.iter_rows():
                if len(row) < min_cols:
                    continue

                sub_id_col  = _get_cell_val(row[CONFIG["COL_SUB_ID"]])
                course_col  = _get_cell_val(row[CONFIG["COL_COURSE"]])
                name_col    = _get_cell_val(row[CONFIG["COL_NAME"]])
                id_col      = _get_cell_val(row[CONFIG["COL_ID"]])
                comment_col = _get_cell_val(row[CONFIG["COL_COMMENT"]])

                comment_cell = row[CONFIG["COL_COMMENT"]]
                fill_color = None
                if comment_cell.fill and comment_cell.fill.patternType == 'solid':
                    fg = comment_cell.fill.start_color
                    if fg.type == 'rgb':
                        fill_color = fg.rgb
                    if fill_color and fill_color not in {'00000000', 'FFFFFFFF', '00FFFFFF'}:
                        pass
                    else:
                        fill_color = None

                if not id_col.strip():
                    if not name_col.strip():
                        continue

                if target_year and not course_col.strip().startswith(str(target_year)):
                    continue

                try:
                    sub_id_val = float(sub_id_col)
                except ValueError:
                    sub_id_val = 0.0

                norm_id = id_col.strip().lower()

                if norm_id in attendance_map_name:
                    final_name = attendance_map_name[norm_id]
                else:
                    c_name_norm = re.sub(r'\s+', '', unicodedata.normalize('NFKC', name_col.strip()))

                    # O(1) exact lookup first; linear scan only as fallback
                    found_id = attendance_name_map.get(c_name_norm)
                    if found_id is None:
                        for att_name_norm, att_id in attendance_name_map.items():
                            if c_name_norm.startswith(att_name_norm):
                                found_id = att_id
                                break

                    if found_id:
                        norm_id = found_id
                        final_name = attendance_map_name[found_id]
                    else:
                        final_name = name_col.strip()

                entry = {
                    'SubmissionID': sub_id_val,
                    'Name': final_name,
                    'ID': id_col.strip(),
                    'NormID': norm_id,
                    'Date': date_str,
                    'Comment': comment_col.strip()
                }
                if fill_color:
                    entry['Color'] = fill_color
                all_data.append(entry)

        except Exception as e:
            print(f"Error processing {file_path}: {e}")
            return False, f"Error processing {os.path.basename(file_path)}: {str(e)}"

    if not all_data and not attendance_ids_order:
        return False, "No data found."

    # ──────────────────────────────────────────────────────────────────────────
    # Deduplicate and build lookups
    # ──────────────────────────────────────────────────────────────────────────
    full_df = pd.DataFrame(all_data)

    if not full_df.empty:
        full_df = full_df.sort_values(by=['NormID', 'Date', 'SubmissionID'])
        full_df = full_df.drop_duplicates(subset=['NormID', 'Date'], keep='last')

    # Color map: (NormID, Date) -> hex color string
    color_map = {}
    if not full_df.empty and 'Color' in full_df.columns:
        colored = full_df[full_df['Color'].notna()]
        color_map = {(r.NormID, r.Date): r.Color for r in colored.itertuples(index=False)}

    # Sorted list of all date column names
    dates = sorted(full_df['Date'].unique().tolist()) if not full_df.empty else []

    # Flat comment lookup: NormID -> {date: comment}
    # Also build student_info to preserve names/IDs for students not in attendance.
    comment_lookup = {}
    student_info = {}   # NormID -> {'name': ..., 'id': ...}  (from comment sheets)
    if not full_df.empty:
        for r in full_df.itertuples(index=False):
            if r.NormID not in comment_lookup:
                comment_lookup[r.NormID] = {}
            comment_lookup[r.NormID][r.Date] = r.Comment
            if r.NormID not in student_info:
                student_info[r.NormID] = {'name': r.Name, 'id': r.ID}

    # ──────────────────────────────────────────────────────────────────────────
    # Write output
    # ──────────────────────────────────────────────────────────────────────────
    print(f"Saving summary to {output_file}")
    try:
        if att_rows_all is not None and attendance_ids_order:
            wb = _build_combined_workbook(
                att_rows_all, comment_lookup, color_map, dates,
                attendance_map_name, attendance_map_original_id,
                student_info=student_info,
                fmt_store=fmt_store,
            )
        else:
            wb = _build_pivot_workbook(
                full_df, dates, color_map,
                attendance_ids_order, attendance_map_name, attendance_map_original_id,
            )

        wb.save(output_file)
        print("Done.")
        return True, f"Saved to {os.path.basename(output_file)}"
    except PermissionError:
        return False, f"Permission denied: {output_file}. Close it and try again."
    except Exception as e:
        return False, f"Error saving file: {str(e)}"


# ──────────────────────────────────────────────────────────────────────────────
# Output builders
# ──────────────────────────────────────────────────────────────────────────────

def _apply_fill(cell, val, norm_id, date, color_map, fill_unanswered, fill_cache):
    """Apply background fill to a comment cell."""
    if val == "未回答":
        cell.fill = fill_unanswered
    elif norm_id:
        hex_c = color_map.get((norm_id, date))
        if hex_c:
            if hex_c not in fill_cache:
                fill_cache[hex_c] = PatternFill(start_color=hex_c, end_color=hex_c, fill_type="solid")
            cell.fill = fill_cache[hex_c]


def _build_combined_workbook(att_rows, comment_lookup, color_map, dates,
                              attendance_map_name, attendance_map_original_id,
                              student_info=None, fmt_store=None):
    """
    Build a workbook that reproduces the attendance sheet (all columns) on the
    left and appends comment columns (Name, ID, one column per date) on the
    right, with each student row aligned.

    Accepts att_rows as a plain list-of-lists so it works identically whether
    the source was an old .xls file (read via xlrd) or a modern .xlsx file
    (read via openpyxl).  The output is always a fresh .xlsx workbook.
    """
    skip_rows = CONFIG["ATT_SKIP_ROWS"]   # first skip_rows rows are header/metadata
    id_0      = CONFIG["ATT_COL_ID"]      # 0-indexed column for student ID
    name_0    = CONFIG["ATT_COL_NAME"]    # 0-indexed column for student name

    # Number of columns in the attendance sheet
    att_col_count = max((len(r) for r in att_rows), default=0)

    # Comment columns start immediately after attendance columns
    col_name  = att_col_count + 1          # 1-indexed for openpyxl
    col_id    = att_col_count + 2
    col_dates = att_col_count + 3          # first date column (1-indexed)

    wb = Workbook()
    ws = wb.active

    fill_unanswered = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    fill_cache = {}
    col_widths = {}

    def _upw(col, val):
        w = len(str(val)) if val is not None else 0
        if col_widths.get(col, 0) < w:
            col_widths[col] = w

    # ── Write all rows from the attendance sheet ─────────────────────────────
    for excel_row_1idx, row in enumerate(att_rows, 1):
        for ci_0idx, val in enumerate(row):
            if val is not None:
                ws.cell(row=excel_row_1idx, column=ci_0idx + 1, value=val)
                _upw(ci_0idx + 1, val)

    # ── Find the main header row (most non-None values in the header section) ─
    # This matches comment headers with the attendance column-name row rather
    # than placing them on a supplementary sub-header row.
    comment_header_row = skip_rows  # safe fallback
    max_non_none = -1
    for r_0 in range(skip_rows):
        count = sum(1 for v in att_rows[r_0] if v is not None)
        if count > max_non_none:
            max_non_none = count
            comment_header_row = r_0 + 1  # convert to 1-indexed Excel row

    # ── Write comment column headers in the main header row ─────────────────
    ws.cell(row=comment_header_row, column=col_name, value="Name")
    ws.cell(row=comment_header_row, column=col_id,   value="ID")
    _upw(col_name, "Name")
    _upw(col_id,   "ID")
    for i, date in enumerate(dates):
        col = col_dates + i
        ws.cell(row=comment_header_row, column=col, value=date)
        _upw(col, date)

    # ── Write comment data for each student row ───────────────────────────────
    seen_nids = set()

    for row_0idx, row in enumerate(att_rows[skip_rows:], skip_rows):
        excel_row = row_0idx + 1   # convert to 1-indexed

        id_raw = row[id_0] if len(row) > id_0 else None
        if id_raw is None:
            continue
        id_str  = _safe_str(id_raw)
        norm_id = id_str.lower()
        if not norm_id or norm_id in _HEADER_IDS:
            continue
        seen_nids.add(norm_id)

        name_str = _safe_str(row[name_0]) if len(row) > name_0 else ''

        ws.cell(row=excel_row, column=col_name, value=name_str)
        ws.cell(row=excel_row, column=col_id,   value=id_str)
        _upw(col_name, name_str)
        _upw(col_id,   id_str)

        student_comments = comment_lookup.get(norm_id, {})
        for i, date in enumerate(dates):
            col      = col_dates + i
            val      = student_comments.get(date, "未回答")
            cell_out = ws.cell(row=excel_row, column=col, value=val)
            _upw(col, val)
            _apply_fill(cell_out, val, norm_id, date, color_map, fill_unanswered, fill_cache)

    # ── Extra rows: students who submitted comments but aren't in attendance ──
    extra_nids = [nid for nid in comment_lookup if nid not in seen_nids]
    if extra_nids:
        next_row = ws.max_row + 1
        _si = student_info or {}
        for nid in extra_nids:
            # Prefer attendance map values; fall back to names from comment sheets
            orig_id   = (attendance_map_original_id.get(nid)
                         or _si.get(nid, {}).get('id', nid))
            orig_name = (attendance_map_name.get(nid)
                         or _si.get(nid, {}).get('name', ''))
            ws.cell(row=next_row, column=col_name, value=orig_name)
            ws.cell(row=next_row, column=col_id,   value=orig_id)
            for i, date in enumerate(dates):
                col      = col_dates + i
                val      = comment_lookup[nid].get(date, "未回答")
                cell_out = ws.cell(row=next_row, column=col, value=val)
                _apply_fill(cell_out, val, nid, date, color_map, fill_unanswered, fill_cache)
            next_row += 1

    # ── Apply .xls formatting to the attendance area ─────────────────────────
    if fmt_store:
        xf_formats   = fmt_store.get('xf', {})
        cells_matrix = fmt_store.get('cl', [])
        fmt_caches   = ({}, {}, {}, {})  # fill, font, border, align

        for r_0idx in range(len(att_rows)):
            excel_row = r_0idx + 1
            row_len   = len(att_rows[r_0idx])
            if r_0idx >= len(cells_matrix):
                break
            for c_0idx in range(min(row_len, len(cells_matrix[r_0idx]))):
                xf_idx = cells_matrix[r_0idx][c_0idx]
                if xf_idx is None:
                    continue
                fmt = xf_formats.get(str(xf_idx))
                if fmt:
                    _apply_att_cell_fmt(
                        ws.cell(row=excel_row, column=c_0idx + 1),
                        fmt, fmt_caches,
                    )

        # Row heights
        for r_str, height_pt in fmt_store.get('rh', {}).items():
            ws.row_dimensions[int(r_str) + 1].height = height_pt

        # Merged cells (attendance area only)
        for mc in fmt_store.get('mg', []):
            rlo, rhi, clo, chi = mc
            # Only merge within the attendance columns
            if chi <= att_col_count:
                try:
                    ws.merge_cells(
                        start_row=rlo + 1,    end_row=rhi,
                        start_column=clo + 1, end_column=chi,
                    )
                except Exception:
                    pass  # ignore duplicate/invalid merges

    # ── Column widths ─────────────────────────────────────────────────────────
    if fmt_store:
        # Attendance columns: use original .xls widths
        for c_str, width in fmt_store.get('cw', {}).items():
            c_1idx = int(c_str) + 1
            if c_1idx <= att_col_count:
                ws.column_dimensions[
                    ws.cell(row=1, column=c_1idx).column_letter
                ].width = width
        # Comment columns: use content-based widths
        for col, width in col_widths.items():
            if col > att_col_count:
                ws.column_dimensions[
                    ws.cell(row=1, column=col).column_letter
                ].width = min(width + 2, 50)
    else:
        # No formatting info: use content-based widths for all columns
        for col, width in col_widths.items():
            ws.column_dimensions[
                ws.cell(row=1, column=col).column_letter
            ].width = min(width + 2, 50)

    return wb


def _build_pivot_workbook(full_df, dates, color_map,
                           attendance_ids_order, attendance_map_name,
                           attendance_map_original_id):
    """
    Build a standalone pivot workbook (used when no attendance file is provided).
    Rows: one per student.  Columns: Name, ID, then one column per date.
    """
    if not full_df.empty:
        pivot_df = full_df.pivot_table(
            index=['NormID'],
            columns='Date',
            values='Comment',
            aggfunc='last',
        )
    else:
        pivot_df = pd.DataFrame()

    if attendance_ids_order:
        existing_ids       = pivot_df.index.tolist() if not pivot_df.empty else []
        attendance_ids_set = set(attendance_ids_order)
        extra_ids          = [i for i in existing_ids if i not in attendance_ids_set]
        full_order_ids     = attendance_ids_order + extra_ids

        pivot_df = pivot_df.reindex(full_order_ids) if not pivot_df.empty else pd.DataFrame(index=full_order_ids)
        pivot_df.index.name = 'NormID'
        pivot_df = pivot_df.reset_index()

        if not full_df.empty:
            extras_map = (full_df[['NormID', 'Name', 'ID']]
                          .drop_duplicates('NormID')
                          .set_index('NormID')[['Name', 'ID']]
                          .to_dict('index'))
        else:
            extras_map = {}

        final_names, final_ids = [], []
        for nid in pivot_df['NormID']:
            final_names.append(
                attendance_map_name.get(nid) or
                (extras_map[nid]['Name'] if nid in extras_map else "Unknown")
            )
            final_ids.append(
                attendance_map_original_id.get(nid) or
                (extras_map[nid]['ID'] if nid in extras_map else nid)
            )

        pivot_df.insert(0, 'Name', final_names)
        pivot_df.insert(1, 'ID',   final_ids)

    else:
        if not full_df.empty:
            pivot_df = pivot_df.reset_index()
            meta_df  = full_df[['NormID', 'Name', 'ID']].drop_duplicates(subset=['NormID'])
            pivot_df = pd.merge(pivot_df, meta_df, on='NormID', how='left')
        else:
            raise ValueError("No data found.")

    pivot_df = pivot_df.fillna("未回答")

    cols        = list(pivot_df.columns)
    date_cols   = sorted(c for c in cols if c not in {'Name', 'ID', 'NormID'})
    output_cols = ['Name', 'ID'] + date_cols

    save_df   = pivot_df[output_cols]
    norm_ids  = pivot_df['NormID'].tolist() if 'NormID' in pivot_df.columns else [None] * len(pivot_df)
    headers   = list(save_df.columns)
    save_vals = save_df.values.tolist()

    wb = Workbook()
    ws = wb.active

    col_widths = [len(str(h)) for h in headers]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)

    fill_unanswered = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    fill_cache = {}

    for ri, (row_data, norm_id) in enumerate(zip(save_vals, norm_ids), 2):
        for meta_col in range(2):
            val = row_data[meta_col]
            ws.cell(row=ri, column=meta_col + 1, value=val)
            col_widths[meta_col] = max(col_widths[meta_col], len(str(val)) if val is not None else 0)

        for ci_offset, (date_col, val) in enumerate(zip(date_cols, row_data[2:])):
            ci   = 3 + ci_offset
            cell = ws.cell(row=ri, column=ci, value=val)
            col_widths[ci - 1] = max(col_widths[ci - 1], len(str(val)) if val is not None else 0)
            _apply_fill(cell, val, norm_id, date_col, color_map, fill_unanswered, fill_cache)

    for ci, width in enumerate(col_widths, 1):
        letter = ws.cell(row=1, column=ci).column_letter
        ws.column_dimensions[letter].width = min(width + 2, 50)

    return wb
